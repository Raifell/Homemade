import openpyxl
from openpyxl.styles import Alignment, Font, Side, Border, PatternFill

head = ['Year', 'Month', 'Category', 'Revenue', 'Cost']
values = [
    [2014, 'January', 'Books', '$37,161', '$28,271'],
    [2014, 'January', 'Electronics', '$354,333', '$283,164'],
    [2014, 'January', 'Movies', '$57,979', '$52,786'],
    [2014, 'January', 'Music', '$52,751', '$48,829'],
    [2014, 'February', 'Books', '$44,810', '$34,133'],
    [2014, 'February', 'Electronics', '$424,982', '$339,971'],
    [2014, 'February', 'Movies', '$71,786', '$65,274'],
    [2014, 'February', 'Music', '$68,478', '$63,389'],
    [2014, 'March', 'Books', '$42,698', '$32,509'],
    [2014, 'March', 'Electronics', '$397,289', '$317,259'],
    [2014, 'March', 'Movies', '$66,886', '$60,913'],
    [2014, 'March', 'Music', '$61,478', '$58,389'],
]

aw = openpyxl.Workbook()
b = aw.active
b.append(head)
for x in values:
    b.append(x)

b.column_dimensions['A'].width = 10
b.column_dimensions['B'].width = 15
b.column_dimensions['C'].width = 15
b.column_dimensions['D'].width = 15
b.column_dimensions['E'].width = 15

for x in range(len(values) + 2):
    b.row_dimensions[x].height = 20

for x in range(len(head)):
    for z in range(1, len(values) + 2):
        cell = b[z][x]
        tolt_1 = Side(border_style="medium", color='FFFFFF')
        tolt_2 = Side(border_style="medium", color='8D8D8D')
        if z == 1:
            b[z][x].alignment = Alignment(horizontal='center', vertical='center')
            b[z][x].font = Font(size=12, color='FFFFFF', bold=True)
            cell.border = Border(bottom=tolt_2, top=tolt_2, left=tolt_2, right=tolt_2)
            cell.fill = PatternFill(fill_type='solid', start_color='363636')
        else:
            if x == 0:
                b[z][x].alignment = Alignment(horizontal='center', vertical='center')
                b[z][x].font = Font(size=12, underline='single', bold=True)
                cell.fill = PatternFill(fill_type='solid', start_color='D3D3D3')
                if z == len(values) + 1:
                    cell.border = Border(bottom=tolt_2, left=tolt_2, right=tolt_1)
                else:
                    cell.border = Border(bottom=tolt_1, left=tolt_2, right=tolt_1)
            elif x == 1:
                b[z][x].alignment = Alignment(horizontal='center', vertical='center')
                b[z][x].font = Font(size=12, underline='single', bold=True)
                cell.fill = PatternFill(fill_type='solid', start_color='D3D3D3')
                if z == len(values) + 1:
                    cell.border = Border(bottom=tolt_2, left=tolt_1, right=tolt_1)
                else:
                    cell.border = Border(bottom=tolt_1, left=tolt_1, right=tolt_1)
            elif x == 2:
                b[z][x].alignment = Alignment(horizontal='left', vertical='center')
                b[z][x].font = Font(size=12, underline='single', bold=True)
                cell.fill = PatternFill(fill_type='solid', start_color='D3D3D3')
                if z == len(values) + 1:
                    cell.border = Border(bottom=tolt_2, left=tolt_1, right=tolt_1)
                else:
                    cell.border = Border(bottom=tolt_1, left=tolt_1, right=tolt_1)
            else:
                b[z][x].alignment = Alignment(horizontal='center', vertical='center')
                b[z][x].font = Font(size=12)
                cell.border = Border(bottom=tolt_2, top=tolt_2, left=tolt_2, right=tolt_2)
                cell.fill = PatternFill(fill_type='solid', start_color='DCEAF9')


aw.save('example.xlsx')

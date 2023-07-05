# Задание №1
# а) Прочитайте из трёх excel файлов (заранее создайте их, внутри 1111, 2222, 3333).
# б) Отсортируйте полученную матрицу в порядке убывания.
# в) Запишите это в один файл, изменив шрифт и обернув в границы.

import openpyxl as oxl
from openpyxl.styles import Alignment, Font, Border, Side



class Task1:
    def __init__(self):
        self.columns = ['A', 'B', 'C', 'D', 'E']
        self.values_1 = [[1, 2, 3, 4, 5], 'table_one.xlsx']
        self.values_2 = [[2, 3, 4, 5, 6], 'table_two.xlsx']
        self.values_3 = [[3, 4, 5, 6, 7], 'table_three.xlsx']
        self.count = []

    @staticmethod
    def create_start(exlist, name):
        aw = oxl.Workbook()
        b = aw.active
        b.append(exlist)
        aw.save(name)

    @staticmethod
    def append_v(count, path):
        wb_obj = oxl.load_workbook(path)
        sheet_obj = wb_obj.active
        exlist = []
        for i in range(1, 6):
            s = sheet_obj.cell(row=1, column=i)
            exlist.append(s.value)
            exlist.sort(reverse=True)
        count.append(exlist)
        return count

    @staticmethod
    def create_result(count, columns):
        aw = oxl.Workbook()
        b = aw.active
        for x in count:
            b.append(x)

        for i in columns:
            for j in range(1, len(count) + 1):
                b[i + str(j)].alignment = Alignment(horizontal='center')
                b[i + str(j)].font = Font(size=12, bold=True, color="FF0000")
                cell = b[i + str(j)]
                tolt = Side(border_style="medium", color="0000ff")
                cell.border = Border(top=tolt, bottom=tolt, left=tolt, right=tolt)

        aw.save('result.xlsx')


a = Task1()
a.create_start(a.values_1[0], a.values_1[1])
a.create_start(a.values_2[0], a.values_2[1])
a.create_start(a.values_3[0], a.values_3[1])

a.append_v(a.count, a.values_1[1])
a.append_v(a.count, a.values_2[1])
a.append_v(a.count, a.values_3[1])

a.create_result(a.count, a.columns)
